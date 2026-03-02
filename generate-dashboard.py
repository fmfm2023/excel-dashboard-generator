"""
Dashboard Excel — Modern BI v3.0
Expert Data Analyst & Senior BI Designer
Style : Modern Infographic, fond blanc, Power BI
"""
import io, os, logging, base64, math, re
from datetime import datetime
import pandas as pd
import numpy as np
from flask import Flask, request, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger(__name__)
app = Flask(__name__)

# ══════════════════════════════════════════════════════════════════════════════
# PALETTE MODERN INFOGRAPHIC (fond blanc, accents vifs)
# ══════════════════════════════════════════════════════════════════════════════
W = {
    # Fond & Neutres
    'white':      'FFFFFF',
    'bg':         'FAFBFE',   # fond très légèrement bleu
    'card':       'FFFFFF',   # fond carte = blanc pur
    'card_border':'E8ECF8',   # bordure carte
    'sep':        'E5E7EB',   # séparateur
    'muted':      'F1F5F9',   # zones muettes
    # Typographie
    'txt_dark':   '0F172A',   # quasi noir
    'txt_mid':    '475569',   # gris foncé
    'txt_light':  '94A3B8',   # gris clair
    'txt_blue':   '3B82F6',   # bleu texte
    # Accents principaux (4 KPI cards)
    'k1':         '4F46E5',   # Indigo   — CA Total
    'k1_l':       'EEF2FF',
    'k2':         '059669',   # Émeraude — Croissance
    'k2_l':       'ECFDF5',
    'k3':         'D97706',   # Ambre    — Panier moy
    'k3_l':       'FFFBEB',
    'k4':         '0EA5E9',   # Sky Blue — Livraison
    'k4_l':       'F0F9FF',
    # Tendances
    'up':         '10B981',   # vert hausse
    'up_l':       'D1FAE5',
    'down':       'EF4444',   # rouge baisse
    'down_l':     'FEE2E2',
    'flat':       '6B7280',   # neutre
    # Graphiques (palette ordonnée)
    'c1':'4F46E5','c2':'059669','c3':'D97706','c4':'0EA5E9','c5':'8B5CF6',
    'c6':'EC4899','c7':'14B8A6','c8':'F97316','c9':'6366F1','c10':'84CC16',
    # Header
    'hdr_bg':     '0F172A',   # fond header = dark navy
    'hdr_accent': '4F46E5',   # ligne accent header
}
CHART_PAL = [W[f'c{i}'] for i in range(1, 11)]

# ══════════════════════════════════════════════════════════════════════════════
# UTILITAIRES STYLE MODERN
# ══════════════════════════════════════════════════════════════════════════════
def _h(c):
    """Normalise hex → 6 chars sans #."""
    return str(c).lstrip('#').upper()[:6].ljust(6,'0')

def fill(c):
    return PatternFill(fgColor=_h(c), fill_type='solid')

def font(sz=10, bold=False, color='0F172A', italic=False, name='Calibri'):
    return Font(name=name, size=sz, bold=bold, color=_h(color), italic=italic)

def align(h='left', v='center', wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def no_border():
    n = Side(style=None)
    return Border(left=n, right=n, top=n, bottom=n)

def thin_border(c='E5E7EB'):
    s = Side(style='thin', color=_h(c))
    return Border(left=s, right=s, top=s, bottom=s)

def left_accent_border(accent_c='4F46E5', side_c='E8ECF8'):
    """Bordure gauche épaisse colorée, simule card style Power BI."""
    thick = Side(style='thick', color=_h(accent_c))
    thin  = Side(style='thin',  color=_h(side_c))
    none  = Side(style=None)
    return Border(left=thick, right=none, top=none, bottom=none)

def full_accent_border(c='E8ECF8'):
    s = Side(style='thin', color=_h(c))
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(c='4F46E5'):
    thick = Side(style='medium', color=_h(c))
    none  = Side(style=None)
    return Border(left=none, right=none, top=none, bottom=thick)

def s(cell, bg=None, fg='0F172A', sz=10, bold=False,
      h='left', v='center', wrap=False, bd=None, nf=None, italic=False, indent=0):
    """Style universel d'une cellule."""
    if bg is not None:
        cell.fill = fill(bg)
    cell.font      = font(sz=sz, bold=bold, color=fg, italic=italic)
    cell.alignment = align(h=h, v=v, wrap=wrap, indent=indent)
    if bd is not None:
        cell.border = bd
    if nf:
        cell.number_format = nf

def mg(ws, r1, c1, r2, c2, val='', bg=None, fg='0F172A',
       sz=11, bold=False, h='left', v='center', wrap=False, bd=None, indent=0):
    """Merge + style."""
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=val)
    if bg is not None:
        for rr in range(r1, r2+1):
            for cc in range(c1, c2+1):
                ws.cell(row=rr, column=cc).fill = fill(bg)
    cell.font      = font(sz=sz, bold=bold, color=fg)
    cell.alignment = align(h=h, v=v, wrap=wrap, indent=indent)
    if bd:
        cell.border = bd
    return cell

def cw(ws, idx, w):
    ws.column_dimensions[get_column_letter(idx)].width = w
def rh(ws, idx, h):
    ws.row_dimensions[idx].height = h

def fnum(v, dec=0):
    if v is None or (isinstance(v, float) and math.isnan(v)): return '—'
    if abs(v) >= 1_000_000: return f'{v/1_000_000:.2f}M'
    if abs(v) >= 1_000:     return f'{v/1_000:.1f}K'
    return f'{v:,.{dec}f}'

def feur(v):
    return fnum(v) + ' €' if v is not None else '—'

def fpct(v, dec=1):
    if v is None or (isinstance(v, float) and math.isnan(v)): return '—'
    return f'{v:.{dec}f}%'

def trend_arrow(v):
    """Retourne (arrow_str, color) selon signe."""
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return ('→', W['flat'])
    if v > 0:  return ('▲', W['up'])
    if v < 0:  return ('▼', W['down'])
    return ('→', W['flat'])

def safe_val(v):
    """Convertit toute valeur invalide pour openpyxl en None (→ cellule vide)."""
    if v is None:
        return None
    # Python float NaN / Inf
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return None
    # NumPy scalaires NaN / Inf
    try:
        if isinstance(v, np.floating) and (np.isnan(v) or np.isinf(v)):
            return None
        if isinstance(v, np.integer):
            return int(v)
    except Exception:
        pass
    # pd.NaT, pd.NA, pd.NAType
    na_str = str(type(v).__name__).lower()
    if 'nat' in na_str or 'natype' in na_str:
        return None
    try:
        if v is pd.NaT or v is pd.NA:
            return None
    except Exception:
        pass
    return v

def safe_float(v, default=0.0):
    """Convertit v en float Python sûr (NaN/Inf/NaT → default)."""
    try:
        f = float(v)
        if math.isnan(f) or math.isinf(f):
            return default
        return f
    except Exception:
        return default

def clean_float_kpi(v):
    """Remplace NaN/Inf par None pour les KPIs optionnels."""
    if v is None:
        return None
    try:
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else f
    except Exception:
        return None

def fnum(v):
    """Formate un nombre entier avec séparateur milliers (sans symbole devise)."""
    try:
        return f"{int(round(float(v))):,}".replace(',', ' ')
    except Exception:
        return str(v) if v is not None else '0'


# ══════════════════════════════════════════════════════════════════════════════
# PART 3 — DATA CLEANING & LOADING
# ══════════════════════════════════════════════════════════════════════════════
def clean_dataframe(df):
    """Dédup, NaN, standardisation dates/devises, colonne Mois/Année."""
    before = len(df)
    df = df.drop_duplicates()
    if before != len(df):
        logger.info(f"Doublons supprimés : {before - len(df)}")
    df.columns = [str(c).strip() for c in df.columns]

    # Nettoyer colonnes monétaires habillées (€, $, %, espaces)
    for col in df.columns:
        if df[col].dtype == object:
            sample = df[col].dropna().astype(str).head(30)
            cleaned = sample.str.replace(r'[€$£%\s]', '', regex=True).str.replace(',', '.')
            if len(cleaned) > 0 and cleaned.str.match(r'^-?\d+\.?\d*$').mean() > 0.7:
                df[col] = pd.to_numeric(
                    df[col].astype(str).str.replace(r'[€$£%\s]', '', regex=True)
                            .str.replace(',', '.'), errors='coerce')

    # Interpoler NaN numériques
    for col in df.select_dtypes(include=[np.number]).columns:
        if df[col].isna().any():
            df[col] = df[col].interpolate(method='linear', limit_direction='both')

    # Détecter & convertir dates
    for col in df.columns:
        if any(kw in col.lower() for kw in ['date', 'jour', 'day', 'time']):
            try:
                df[col] = pd.to_datetime(df[col], dayfirst=True, errors='coerce')
            except Exception:
                pass

    # Créer colonne Mois/Année depuis première date trouvée
    date_cols = df.select_dtypes(include=['datetime64[ns]']).columns.tolist()
    if date_cols:
        df['Mois/Année'] = df[date_cols[0]].dt.to_period('M').astype(str)
    else:
        df['Mois/Année'] = 'N/A'
    return df

def find_header_row(df_test):
    """Détecte si la 1ère ligne est un titre ou une vraie en-tête."""
    if len(df_test) == 0:
        return 0
    n_filled = (df_test.iloc[0].astype(str).str.strip() != '').sum()
    return 1 if n_filled < max(3, len(df_test.columns) * 0.3) else 0


def load_dataframe(file_bytes, filename):
    """Charge CSV ou Excel en DataFrame pandas."""
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else 'xlsx'
    try:
        if ext == 'csv':
            df = None
            for enc in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    df = pd.read_csv(io.BytesIO(file_bytes), encoding=enc,
                                     sep=None, engine='python', dtype=str)
                    if len(df.columns) >= 2:
                        break
                except Exception:
                    continue
            if df is None:
                raise ValueError("Impossible de lire le CSV")
        else:
            df_test = pd.read_excel(io.BytesIO(file_bytes), header=0, dtype=str, nrows=3)
            hdr = find_header_row(df_test)
            df = pd.read_excel(io.BytesIO(file_bytes), header=hdr, dtype=str)

        # Forcer conversion numérique colonne par colonne
        # (pandas 2.2+ : errors='ignore' supprimé → on simule avec coerce + seuil)
        for col in df.columns:
            converted = pd.to_numeric(df[col], errors='coerce')
            if converted.notna().mean() > 0.5:   # >50% valeurs numériques → garder
                df[col] = converted
        df.columns = [str(c).strip() for c in df.columns]
        return df
    except Exception as e:
        logger.error(f"load_dataframe: {e}")
        raise


# ══════════════════════════════════════════════════════════════════════════════
# PART 4 — DÉTECTION DES COLONNES
# ══════════════════════════════════════════════════════════════════════════════
def detect_columns(df):
    """Mappe les colonnes à des rôles sémantiques (FR/EN + pharma/stock)."""
    cm = {}
    low = {c.lower().strip(): c for c in df.columns}
    ROLES = {
        'date':       ['date', 'jour', 'day', 'time', 'période'],
        'peremption': ['peremption', 'péremption', 'expiry', 'expiration', 'expire'],
        'client':     ['client', 'customer', 'acheteur', 'compte'],
        'produit':    ['produit', 'product', 'article', 'item', 'désignation', 'libellé',
                       'medicament', 'médicament', 'nom_med', 'nom med'],
        'categorie':  ['catégorie', 'categorie', 'category', 'famille', 'gamme',
                       'type produit', 'gamme_therapeutique', 'therapeutique'],
        'marque':     ['marque', 'brand', 'fabricant', 'forme', 'laboratoire'],
        'quantite':   ['quantité', 'quantite', 'qte', 'qty', 'quantity', 'volume'],
        'stock':      ['stock_actuel', 'stock actuel', 'stock', 'inventaire', 'disponible'],
        'prix':       ['prix_unitaire', 'prix unitaire', 'prix', 'price', 'tarif', 'coût', 'cout'],
        'objectif':   ['objectif_vente', 'objectif vente', 'objectif', 'target', 'cible',
                       'budget', 'prevision', 'prévision'],
        'realise':    ['ventes_realisees', 'ventes realisees', 'vente_realisee', 'realise',
                       'réalisé', 'realisee', 'achieved', 'actual'],
        'delegue':    ['delegue_medical', 'delegue medical', 'délégué', 'delegue',
                       'representant', 'rep'],
        'remise':     ['remise', 'discount', 'rabais', 'réduction', 'taux remise'],
        'ca':         ['total ttc', 'total ht', 'chiffre', 'montant', 'revenue', 'total', 'amount'],
        'tva':        ['tva', 'vat', 'taxe'],
        'statut':     ['statut', 'status', 'état', 'etat'],
        'vendeur':    ['vendeur', 'seller', 'commercial'],
        'region':     ['région', 'region', 'zone', 'territoire', 'pays', 'ville'],
    }
    for role, kws in ROLES.items():
        for kw in kws:
            for col_l, col_o in low.items():
                if kw in col_l and role not in cm:
                    cm[role] = col_o
                    break

    # Priorité CA : Total TTC > Total HT > autres
    for pk in ['total ttc', 'total ht', 'montant ttc', 'ca total', 'chiffre affaires']:
        for col_l, col_o in low.items():
            if pk in col_l:
                cm['ca'] = col_o
                break
        if 'ca' in cm:
            break

    # Fallback CA : si pas de CA mais 'realise' détecté → utiliser réalisé comme CA
    if 'ca' not in cm and 'realise' in cm:
        cm['ca'] = cm['realise']

    # Fallback vendeur → délégué médical
    if 'vendeur' not in cm and 'delegue' in cm:
        cm['vendeur'] = cm['delegue']

    logger.info(f"Colonnes détectées: {list(cm.keys())}")
    return cm


# ══════════════════════════════════════════════════════════════════════════════
# PART 5 — KPIs AVANCÉS
# ══════════════════════════════════════════════════════════════════════════════
def compute_advanced_kpis(df, cm):
    """Calcule CA, MoM/YoY, panier, statuts, top segments + KPIs pharma/stock."""
    kpis = {'n_rows': len(df)}
    ca_col = cm.get('ca')
    kpis['ca_col'] = ca_col
    ca_s = pd.to_numeric(df[ca_col], errors='coerce') if ca_col and ca_col in df.columns else pd.Series(dtype=float)
    kpis['ca_total'] = float(ca_s.sum()) if len(ca_s) > 0 else 0.0

    qty_col = cm.get('quantite')
    kpis['qty_total'] = float(pd.to_numeric(df[qty_col], errors='coerce').sum()) \
        if qty_col and qty_col in df.columns else float(kpis['n_rows'])

    cli_col = cm.get('client')
    kpis['n_clients'] = int(df[cli_col].nunique()) if cli_col and cli_col in df.columns else 0
    kpis['panier_moy'] = kpis['ca_total'] / max(kpis['n_clients'], 1)

    # Statuts
    stat_col = cm.get('statut')
    if stat_col and stat_col in df.columns:
        vc  = df[stat_col].value_counts()
        n   = len(df)
        POS = ['livré', 'livrée', 'delivered', 'terminé', 'completed', 'payé', 'validé']
        NEG = ['annulé', 'annulée', 'cancelled', 'canceled', 'refusé']
        dlv = sum(v for k, v in vc.items() if any(p in str(k).lower() for p in POS))
        can = sum(v for k, v in vc.items() if any(p in str(k).lower() for p in NEG))
        kpis['taux_livraison']  = dlv / n * 100 if n > 0 else 0
        kpis['taux_annulation'] = can / n * 100 if n > 0 else 0
        kpis['statut_vc'] = {str(k): int(v) for k, v in vc.items()}
    else:
        kpis['taux_livraison'] = kpis['taux_annulation'] = None
        kpis['statut_vc'] = {}

    rem_col = cm.get('remise')
    _rem = float(pd.to_numeric(df[rem_col], errors='coerce').mean()) \
        if rem_col and rem_col in df.columns else None
    kpis['remise_moy'] = clean_float_kpi(_rem)
    tva_col = cm.get('tva')
    _tva = float(pd.to_numeric(df[tva_col], errors='coerce').sum()) \
        if tva_col and tva_col in df.columns else 0.0
    kpis['tva_total'] = clean_float_kpi(_tva) or 0.0

    # Évolution mensuelle
    if 'Mois/Année' in df.columns and ca_col and ca_col in df.columns:
        monthly = (df.groupby('Mois/Année')[ca_col]
                   .apply(lambda x: pd.to_numeric(x, errors='coerce').sum())
                   .sort_index())
        kpis['monthly_ca'] = {str(k): safe_float(v) for k, v in monthly.items()}
        ml = monthly.tolist()
        if len(ml) >= 2 and ml[-2] != 0:
            kpis['mom_growth'] = (ml[-1] - ml[-2]) / abs(ml[-2]) * 100
        else:
            kpis['mom_growth'] = None
        kpis['yoy_growth'] = (ml[-1] - ml[-13]) / abs(ml[-13]) * 100 \
            if len(ml) >= 13 and ml[-13] != 0 else None
    else:
        kpis['monthly_ca'] = {}
        kpis['mom_growth'] = kpis['yoy_growth'] = None

    # Top segments
    def _top(role, val_role='ca', n=8):
        col = cm.get(role)
        vcol = cm.get(val_role) or ca_col
        if col and col in df.columns and vcol and vcol in df.columns:
            raw = (df.groupby(col)[vcol]
                   .apply(lambda x: pd.to_numeric(x, errors='coerce').sum())
                   .sort_values(ascending=False).head(n).to_dict())
            return {str(k): safe_float(v) for k, v in raw.items()}
        return {}

    kpis['top_produits']   = _top('produit')
    kpis['top_categories'] = _top('categorie')
    kpis['top_vendeurs']   = _top('vendeur')
    kpis['top_delegues']   = _top('delegue', 'realise') or _top('delegue')
    for role in ('produit', 'categorie', 'vendeur'):
        col = cm.get(role)
        kpis[f'n_{role}s'] = int(df[col].nunique()) if col and col in df.columns else 0

    # ── KPIs PHARMA / STOCK ──────────────────────────────────────────────────
    stock_col  = cm.get('stock')
    prix_col   = cm.get('prix')
    obj_col    = cm.get('objectif')
    real_col   = cm.get('realise')
    # Ne pas confondre realise == ca (fallback) avec un vrai réalisé distinct
    real_is_ca = (real_col == ca_col)
    delg_col   = cm.get('delegue')
    perm_col   = cm.get('peremption')

    # Stock total + ruptures
    if stock_col and stock_col in df.columns:
        stock_s = pd.to_numeric(df[stock_col], errors='coerce')
        kpis['stock_total']  = float(stock_s.sum())
        kpis['n_ruptures']   = int((stock_s == 0).sum())
        kpis['stock_faible'] = int(((stock_s > 0) & (stock_s < stock_s.mean() * 0.2)).sum()) \
            if stock_s.mean() > 0 else 0
    else:
        kpis['stock_total'] = kpis['n_ruptures'] = kpis['stock_faible'] = None

    # CA Potentiel = sum(stock × prix)
    if stock_col and prix_col and stock_col in df.columns and prix_col in df.columns:
        stk = pd.to_numeric(df[stock_col], errors='coerce').fillna(0)
        prx = pd.to_numeric(df[prix_col],  errors='coerce').fillna(0)
        kpis['ca_potentiel'] = float((stk * prx).sum())
    else:
        kpis['ca_potentiel'] = None

    # Taux réalisation (objectif vs réalisé)
    # Calculer si obj et realise sont deux colonnes DISTINCTES et toutes deux présentes
    if obj_col and real_col and obj_col != real_col \
            and obj_col in df.columns and real_col in df.columns:
        obj_s  = pd.to_numeric(df[obj_col],  errors='coerce').fillna(0)
        real_s = pd.to_numeric(df[real_col], errors='coerce').fillna(0)
        obj_t  = float(obj_s.sum())
        real_t = float(real_s.sum())
        kpis['objectif_total']         = obj_t
        kpis['realise_total']          = real_t
        kpis['taux_realisation_global'] = (real_t / obj_t * 100) if obj_t > 0 else None

        # Performance par délégué
        if delg_col and delg_col in df.columns:
            grp_o = df.groupby(delg_col)[obj_col].apply(
                lambda x: pd.to_numeric(x, errors='coerce').sum())
            grp_r = df.groupby(delg_col)[real_col].apply(
                lambda x: pd.to_numeric(x, errors='coerce').sum())
            perf = []
            for name in grp_o.index:
                o = safe_float(grp_o.get(name, 0))
                r = safe_float(grp_r.get(name, 0))
                t = (r / o * 100) if o > 0 else 0
                perf.append({'delegue': str(name), 'objectif': o, 'realise': r, 'taux': t})
            perf.sort(key=lambda x: x['taux'], reverse=True)
            kpis['perf_delegues'] = perf
        else:
            kpis['perf_delegues'] = []
    else:
        kpis['objectif_total'] = kpis['realise_total'] = None
        kpis['taux_realisation_global'] = None
        kpis['perf_delegues'] = []

    # Péremptions (alertes)
    if perm_col and perm_col in df.columns:
        today = pd.Timestamp.now()
        d90   = today + pd.Timedelta(days=90)
        perm_s = pd.to_datetime(df[perm_col], errors='coerce')
        kpis['n_perimes']          = int((perm_s < today).sum())
        kpis['n_peremption_proche'] = int(((perm_s >= today) & (perm_s <= d90)).sum())
    else:
        kpis['n_perimes'] = kpis['n_peremption_proche'] = None

    logger.info(f"KPIs: CA={kpis['ca_total']:.0f} rows={kpis['n_rows']} "
                f"stock={kpis.get('stock_total')} taux_real={kpis.get('taux_realisation_global')}")
    return kpis


# ══════════════════════════════════════════════════════════════════════════════
# PART 6 — AI INSIGHTS
# ══════════════════════════════════════════════════════════════════════════════
def generate_insights(df, kpis, cm):
    """Génère 3 insights analytiques automatiques."""
    insights = []

    # Insight 1 — taux réalisation ou tendance MoM
    taux_r = kpis.get('taux_realisation_global')
    mom    = kpis.get('mom_growth')
    if taux_r is not None:
        icon = '✅' if taux_r >= 100 else ('⚠️' if taux_r >= 80 else '🔴')
        qual = 'atteint' if taux_r >= 100 else ('presque atteint' if taux_r >= 80 else 'insuffisant')
        insights.append(f"{icon} Taux de réalisation global : {taux_r:.1f}% — objectif {qual}. "
                        f"({fnum(kpis.get('realise_total',0))} / {fnum(kpis.get('objectif_total',0))})")
    elif mom is not None:
        d = 'progression' if mom >= 0 else 'recul'
        e = '📈' if mom >= 0 else '📉'
        insights.append(f"{e} Tendance MoM : {d} de {abs(mom):.1f}% du CA vs mois précédent.")
    elif kpis.get('monthly_ca'):
        ml = list(kpis['monthly_ca'].keys())
        insights.append(f"📊 Données sur {len(ml)} période(s) · de {ml[0]} à {ml[-1]}.")
    else:
        insights.append(f"📦 {kpis['n_rows']} enregistrements analysés.")

    # Insight 2 — stock/ruptures ou top performer
    n_rupt = kpis.get('n_ruptures')
    n_perm = kpis.get('n_peremption_proche')
    ca_pot = kpis.get('ca_potentiel')
    tv     = kpis.get('top_delegues', {}) or kpis.get('top_vendeurs', {})
    tp     = kpis.get('top_produits', {})
    total  = kpis.get('ca_total', 1) or 1
    if n_rupt is not None:
        stk_tot = kpis.get('stock_total', 0) or 0
        insights.append(
            f"📦 Stock total : {fnum(stk_tot)} unités · {n_rupt} rupture(s)"
            + (f" · {n_perm} produit(s) périmant dans 90j" if n_perm else "")
            + (f" · CA potentiel : {feur(ca_pot)}" if ca_pot else "")
        )
    elif tv:
        k, v = list(tv.items())[0]
        pct = v / total * 100
        conc = 'élevée' if pct > 40 else 'équilibrée'
        insights.append(f"🏆 Top délégué/vendeur : {k} · {pct:.1f}% du CA ({feur(v)}) — concentration {conc}.")
    elif tp:
        k, v = list(tp.items())[0]
        insights.append(f"🏆 Top produit : « {str(k)[:25]} » · {v/total*100:.1f}% du CA ({feur(v)}).")

    # Insight 3 — alerte ou opportunité
    ann  = kpis.get('taux_annulation') or 0
    rem  = kpis.get('remise_moy')
    cats = list(kpis.get('top_categories', {}).items())[:2]
    perf = kpis.get('perf_delegues', [])
    if perf:
        top_d    = perf[0]
        bottom_d = perf[-1] if len(perf) > 1 else None
        msg = f"🏆 Meilleur délégué : {top_d['delegue']} ({top_d['taux']:.0f}%)"
        if bottom_d and bottom_d['taux'] < 80:
            msg += f" · À améliorer : {bottom_d['delegue']} ({bottom_d['taux']:.0f}%)"
        insights.append(msg)
    elif ann > 15:
        insights.append(f"⚠️ Taux d'annulation élevé ({ann:.1f}%) — analyser stock & délais pour réduire les pertes.")
    elif rem and rem > 10:
        gain = kpis['ca_total'] * 0.02
        insights.append(f"💡 Remise moy. {rem:.1f}% — réduire de 2pp augmenterait le CA de ~{feur(gain)}.")
    elif len(cats) >= 2:
        pct2 = sum(v for _, v in cats) / total * 100
        insights.append(f"💡 Top 2 catégories ({cats[0][0]}, {cats[1][0]}) = {pct2:.1f}% du CA — levier prioritaire.")
    else:
        insights.append(f"👥 {kpis['n_clients']} clients · panier moyen {feur(kpis.get('panier_moy', 0))}.")

    # Compléter à 3
    while len(insights) < 3:
        insights.append(f"📊 {kpis['n_rows']} enregistrements · {kpis.get('n_produits', 0)} produits.")
    return insights[:3]


# ══════════════════════════════════════════════════════════════════════════════
# PART 7 — DASHBOARD PRINCIPAL (Modern Infographic)
# ══════════════════════════════════════════════════════════════════════════════
def _build_kpi_cards(kpis, cm):
    """Construit 4 cartes KPI dynamiques selon les données disponibles."""
    col_positions = [2, 7, 12, 17]

    # Card 1 (indigo) — Principal chiffre de vente / réalisé
    real_total = kpis.get('realise_total')
    ca_total   = kpis.get('ca_total', 0)
    if real_total and real_total > 0:
        c1 = ('VENTES RÉALISÉES', fnum(real_total), None,
              f"Objectif : {fnum(kpis.get('objectif_total', 0))}",
              W['k1'], W['k1_l'], col_positions[0])
    else:
        c1 = ('CA TOTAL', feur(ca_total), kpis.get('mom_growth'),
              f"{kpis['n_rows']} enregistrements",
              W['k1'], W['k1_l'], col_positions[0])

    # Card 2 (émeraude) — Taux réalisation ou croissance MoM
    taux = kpis.get('taux_realisation_global')
    if taux is not None:
        delta = taux - 100
        c2 = ('TAUX RÉALISATION', f"{taux:.1f}%",
              delta if abs(delta) > 0.1 else None,
              'vs Objectif fixé',
              W['k2'], W['k2_l'], col_positions[1])
    else:
        mom = kpis.get('mom_growth')
        c2 = ('CROISSANCE MoM',
              fpct(mom) if mom is not None else '—',
              mom, 'vs mois précédent',
              W['k2'], W['k2_l'], col_positions[1])

    # Card 3 (ambre) — CA potentiel ou Stock total ou Panier moyen
    ca_pot     = kpis.get('ca_potentiel')
    stock_tot  = kpis.get('stock_total')
    if ca_pot and ca_pot > 0:
        c3 = ('CA POTENTIEL', feur(ca_pot), None,
              'Stock × Prix unitaire',
              W['k3'], W['k3_l'], col_positions[2])
    elif stock_tot and stock_tot > 0:
        c3 = ('STOCK TOTAL', fnum(stock_tot), None,
              f"Ruptures : {kpis.get('n_ruptures', 0)}",
              W['k3'], W['k3_l'], col_positions[2])
    else:
        c3 = ('PANIER MOYEN', feur(kpis.get('panier_moy', 0)), None,
              f"{kpis.get('n_clients', 0)} clients",
              W['k3'], W['k3_l'], col_positions[2])

    # Card 4 (sky) — Ruptures stock ou Taux livraison ou Nb produits
    n_rupt = kpis.get('n_ruptures')
    if n_rupt is not None:
        n_exp = kpis.get('n_peremption_proche') or 0
        c4 = ('RUPTURES STOCK', str(n_rupt),
              -float(n_rupt) if n_rupt > 0 else None,
              f"Péremption <90j : {n_exp}",
              W['k4'], W['k4_l'], col_positions[3])
    elif kpis.get('taux_livraison') is not None:
        c4 = ('TAUX LIVRAISON',
              fpct(kpis.get('taux_livraison')), None,
              f"Annul. : {fpct(kpis.get('taux_annulation'))}",
              W['k4'], W['k4_l'], col_positions[3])
    else:
        n_p = kpis.get('n_produits') or kpis.get('n_categories') or kpis['n_rows']
        c4 = ('NB PRODUITS', str(n_p), None,
              cm.get('produit', 'Références'),
              W['k4'], W['k4_l'], col_positions[3])

    return [c1, c2, c3, c4]


def build_modern_dashboard(wb, df, kpis, cm, insights):
    ws = wb.create_sheet('📊 Dashboard')
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    # ── Colonnes
    cw(ws, 1, 1.5)
    for ci in range(2, 6):   cw(ws, ci, 12)    # B-E card1
    cw(ws, 6, 1.5)
    for ci in range(7, 11):  cw(ws, ci, 12)    # G-J card2
    cw(ws, 11, 1.5)
    for ci in range(12, 16): cw(ws, ci, 12)    # L-O card3
    cw(ws, 16, 1.5)
    for ci in range(17, 21): cw(ws, ci, 12)    # Q-T card4
    cw(ws, 21, 1.5)
    for ci in range(22, 40): cw(ws, ci, 15)    # data zone cachée

    # ── Hauteurs de lignes
    rh(ws, 1, 8);  rh(ws, 2, 35); rh(ws, 3, 8)
    rh(ws, 4, 5);  rh(ws, 5, 28); rh(ws, 6, 18); rh(ws, 7, 18); rh(ws, 8, 5)
    rh(ws, 9, 22)
    for r in range(10, 33): rh(ws, r, 14)
    rh(ws, 33, 22)
    for r in range(34, 56): rh(ws, r, 14)
    rh(ws, 56, 22)
    for r in range(57, 72): rh(ws, r, 20)
    rh(ws, 72, 22)
    for r in range(73, 84): rh(ws, r, 22)
    rh(ws, 84, 14)

    # ── Fond général
    for r in range(1, 87):
        for c in range(1, 30):
            ws.cell(r, c).fill = fill(W['bg'])

    # ── HEADER (rows 1-2)
    for r in [1, 2]:
        for c in range(1, 22):
            ws.cell(r, c).fill = fill(W['hdr_bg'])
    mg(ws, 2, 2, 2, 14, '📊  DASHBOARD ANALYTIQUE — MODERN BI v3',
       bg=W['hdr_bg'], fg=W['white'], sz=18, bold=True, h='left', v='center')
    mg(ws, 2, 15, 2, 20,
       f"{datetime.now().strftime('%d/%m/%Y')} · {kpis['n_rows']} enregistrements",
       bg=W['hdr_bg'], fg=W['txt_light'], sz=9, h='right', v='center')
    for c in range(1, 22):
        ws.cell(2, c).border = bottom_border(W['hdr_accent'])

    # ── KPI CARDS (dynamiques)
    cards = _build_kpi_cards(kpis, cm)
    for lbl, val, growth, sub, acc, light, col in cards:
        for r in range(4, 9):
            for c in range(col, col + 4):
                ws.cell(r, c).fill = fill(W['card'])
                ws.cell(r, c).border = full_accent_border(W['card_border'])
        for r in range(4, 9):
            ws.cell(r, col).fill = fill(acc)
        ws.merge_cells(start_row=5, start_column=col+1, end_row=5, end_column=col+3)
        cell_v = ws.cell(5, col+1, val)
        s(cell_v, bg=W['card'], fg=acc, sz=20, bold=True, h='left', v='center')
        ws.merge_cells(start_row=6, start_column=col+1, end_row=6, end_column=col+3)
        cell_l = ws.cell(6, col+1, lbl)
        s(cell_l, bg=W['card'], fg=W['txt_mid'], sz=9, bold=True, h='left', v='center')
        if growth is not None:
            arr, arr_c = trend_arrow(growth)
            sub_txt = f"{arr} {fpct(abs(growth))}  ·  {sub}"
        else:
            arr_c   = W['txt_light']
            sub_txt = sub
        ws.merge_cells(start_row=7, start_column=col+1, end_row=7, end_column=col+3)
        cell_s = ws.cell(7, col+1, sub_txt)
        s(cell_s, bg=W['card'], fg=arr_c, sz=9, h='left', v='center')

    # ── SECTION GRAPHIQUES
    mg(ws, 9, 2, 9, 20, '  📈  ANALYSE GRAPHIQUE',
       bg=W['bg'], fg=W['txt_dark'], sz=11, bold=True, h='left', v='center')

    # ── DONNÉES pour graphiques (colonnes cachées 22+)
    DC = 22   # DATA_COL_START

    # --- Graphique 1 : Bar par catégorie/gamme (DC, DC+1)
    cat_col = cm.get('categorie') or cm.get('produit')
    val_col = cm.get('realise') or cm.get('ca')
    chart1_data = []
    if cat_col and val_col and cat_col in df.columns and val_col in df.columns:
        grp = (df.groupby(cat_col)[val_col]
               .apply(lambda x: pd.to_numeric(x, errors='coerce').sum())
               .sort_values(ascending=False).head(8))
        chart1_data = [(str(k)[:22], safe_float(v)) for k, v in grp.items()]
    ws.cell(10, DC, 'Segment');  ws.cell(10, DC+1, 'Valeur')
    for i, (lb, vl) in enumerate(chart1_data, 1):
        ws.cell(10+i, DC, lb);  ws.cell(10+i, DC+1, vl)
    n_c1 = len(chart1_data)
    if n_c1 >= 2:
        bc = BarChart()
        bc.type = 'col';  bc.style = 10;  bc.grouping = 'clustered'
        bc.title = f"Ventes par {cat_col or 'Segment'}"
        cats_r = Reference(ws, min_col=DC,   min_row=11, max_row=10+n_c1)
        data_r = Reference(ws, min_col=DC+1, min_row=10, max_row=10+n_c1)
        bc.add_data(data_r, titles_from_data=True)
        bc.set_categories(cats_r)
        bc.series[0].graphicalProperties.solidFill = W['k1']
        bc.series[0].graphicalProperties.line.solidFill = W['k1']
        bc.width = 12.5;  bc.height = 9
        ws.add_chart(bc, 'B10')

    # --- Graphique 2 : Line Objectif vs Réalisé (DC+2, DC+3, DC+4)
    obj_col  = cm.get('objectif')
    real_col = cm.get('realise')
    real_is_ca = (real_col == cm.get('ca'))
    chart2_data = []
    line_s2 = None
    if obj_col and real_col and obj_col != real_col \
            and obj_col in df.columns and real_col in df.columns and cat_col in df.columns:
        grp_o = (df.groupby(cat_col)[obj_col]
                 .apply(lambda x: pd.to_numeric(x, errors='coerce').sum())
                 .sort_values(ascending=False).head(8))
        grp_r = (df.groupby(cat_col)[real_col]
                 .apply(lambda x: pd.to_numeric(x, errors='coerce').sum())
                 .reindex(grp_o.index))
        chart2_data = [(str(k)[:22], safe_float(vo), safe_float(vr))
                       for (k, vo), vr in zip(grp_o.items(), grp_r.values)]
        line_title = f"Objectif vs Réalisé par {cat_col}"
        line_s1 = 'Objectif';  line_s2 = 'Réalisé'
    else:
        monthly = kpis.get('monthly_ca', {})
        sm = sorted(monthly.items())[:18]
        chart2_data = [(str(m), safe_float(v), None) for m, v in sm]
        line_title = "Évolution CA Mensuel";  line_s1 = 'CA'
    ws.cell(10, DC+2, 'Label');  ws.cell(10, DC+3, line_s1)
    if line_s2:
        ws.cell(10, DC+4, line_s2)
    for i, rd in enumerate(chart2_data, 1):
        ws.cell(10+i, DC+2, rd[0]);  ws.cell(10+i, DC+3, rd[1])
        if rd[2] is not None:
            ws.cell(10+i, DC+4, rd[2])
    n_c2 = len(chart2_data)
    if n_c2 >= 2:
        lc = LineChart();  lc.style = 10;  lc.title = line_title
        cats_r = Reference(ws, min_col=DC+2, min_row=11, max_row=10+n_c2)
        max_c  = DC+4 if (line_s2 and any(rd[2] is not None for rd in chart2_data)) else DC+3
        data_r = Reference(ws, min_col=DC+3, min_row=10, max_row=10+n_c2, max_col=max_c)
        lc.add_data(data_r, titles_from_data=True)
        lc.set_categories(cats_r)
        for i, (clr, wt) in enumerate([(W['k1'], 2.5), (W['k2'], 2.0)]):
            if i < len(lc.series):
                lc.series[i].graphicalProperties.line.solidFill = clr
                lc.series[i].graphicalProperties.line.width = int(wt * 12700)
                lc.series[i].marker.symbol = 'circle'
                lc.series[i].marker.size   = 5
        lc.width = 12.5;  lc.height = 9
        ws.add_chart(lc, 'L10')

    # ── SECTION RÉPARTITION & TOP
    mg(ws, 33, 2, 33, 20, '  🎯  RÉPARTITION & TOP PERFORMERS',
       bg=W['bg'], fg=W['txt_dark'], sz=11, bold=True, h='left', v='center')

    # --- Graphique 3 : Pie/Doughnut répartition (DC+5, DC+6)
    top_seg = (list(kpis.get('top_categories', {}).items())[:7]
               or list(kpis.get('top_produits', {}).items())[:7])
    if not top_seg and cat_col and val_col \
            and cat_col in df.columns and val_col in df.columns:
        grp = (df.groupby(cat_col)[val_col]
               .apply(lambda x: pd.to_numeric(x, errors='coerce').sum())
               .sort_values(ascending=False).head(7))
        top_seg = [(str(k)[:22], safe_float(v)) for k, v in grp.items()]
    n_seg = len(top_seg)
    ws.cell(34, DC+5, 'Segment');  ws.cell(34, DC+6, 'Valeur')
    for i, (lb, vl) in enumerate(top_seg, 1):
        ws.cell(34+i, DC+5, str(lb)[:22]);  ws.cell(34+i, DC+6, safe_float(vl))
    if n_seg >= 2:
        pie = PieChart();  pie.style = 10;  pie.title = "Répartition par Segment"
        cats_p = Reference(ws, min_col=DC+5, min_row=35, max_row=34+n_seg)
        data_p = Reference(ws, min_col=DC+6, min_row=34, max_row=34+n_seg)
        pie.add_data(data_p, titles_from_data=True)
        pie.set_categories(cats_p)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True
        for i in range(n_seg):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = CHART_PAL[i % len(CHART_PAL)]
            pie.series[0].dPt.append(pt)
        pie.width = 12.5;  pie.height = 9
        ws.add_chart(pie, 'B34')

    # --- Graphique 4 : Top-10 horizontal bar (DC+7, DC+8)
    top_items = (list(kpis.get('top_delegues', {}).items())[:10]
                 or list(kpis.get('top_vendeurs', {}).items())[:10]
                 or list(kpis.get('top_produits', {}).items())[:10])
    top_label = (cm.get('delegue') or cm.get('vendeur') or cm.get('produit') or 'Top')
    n_top = len(top_items)
    ws.cell(34, DC+7, str(top_label)[:14]);  ws.cell(34, DC+8, 'Valeur')
    for i, (lb, vl) in enumerate(top_items, 1):
        ws.cell(34+i, DC+7, str(lb)[:25]);  ws.cell(34+i, DC+8, safe_float(vl))
    if n_top >= 2:
        hb = BarChart();  hb.type = 'bar';  hb.style = 10;  hb.grouping = 'clustered'
        hb.title = f"Top {top_label}"
        cats_h = Reference(ws, min_col=DC+7, min_row=35, max_row=34+n_top)
        data_h = Reference(ws, min_col=DC+8, min_row=34, max_row=34+n_top)
        hb.add_data(data_h, titles_from_data=True)
        hb.set_categories(cats_h)
        hb.series[0].graphicalProperties.solidFill = W['k4']
        hb.series[0].graphicalProperties.line.solidFill = W['k4']
        hb.width = 12.5;  hb.height = 9
        ws.add_chart(hb, 'L34')

    # ── SECTION PERFORMANCE DÉLÉGUÉS
    perf_data = kpis.get('perf_delegues', [])
    if perf_data:
        mg(ws, 56, 2, 56, 20, '  📊  PERFORMANCE DÉLÉGUÉS — TAUX DE RÉALISATION',
           bg=W['bg'], fg=W['txt_dark'], sz=11, bold=True, h='left', v='center')
        hdrs_p = ['Délégué', 'Objectif', 'Réalisé', 'Taux %', 'Progression']
        h_cols = [2, 6, 9, 12, 14]
        for h_txt, ci in zip(hdrs_p, h_cols):
            cell = ws.cell(57, ci, h_txt)
            s(cell, bg=W['hdr_bg'], fg=W['white'], sz=10, bold=True, h='center', v='center')
        for idx, pd_row in enumerate(perf_data[:10]):
            r_p  = 58 + idx
            taux = pd_row.get('taux', 0) or 0
            rbg  = W['k2_l'] if taux >= 100 else (W['k3_l'] if taux >= 80 else W['down_l'])
            bar_len  = min(int(taux / 10), 10)
            prog_bar = '█' * bar_len + '░' * (10 - bar_len)
            row_vals = [
                (2,  str(pd_row.get('delegue', ''))[:30], W['txt_dark'], True),
                (6,  fnum(pd_row.get('objectif', 0)),      W['txt_mid'],  False),
                (9,  fnum(pd_row.get('realise', 0)),
                     W['k2'] if taux >= 100 else W['k1'],  True),
                (12, f"{taux:.1f}%",
                     W['k2'] if taux >= 100 else W['down'], True),
                (14, prog_bar, W['k2'] if taux >= 100 else W['k3'], False),
            ]
            for ci, cv, fg_c, bold_c in row_vals:
                cell = ws.cell(r_p, ci, cv)
                s(cell, bg=rbg, fg=fg_c, sz=10, bold=bold_c,
                  h='left' if ci == 2 else 'center')
                cell.border = thin_border(W['sep'])

    # ── AI INSIGHTS
    mg(ws, 72, 2, 72, 20, '  🤖  AI INSIGHTS',
       bg=W['hdr_bg'], fg=W['white'], sz=11, bold=True, h='left', v='center')
    ins_bgs  = [W['k1_l'], W['k2_l'], W['k3_l']]
    ins_accs = [W['k1'],   W['k2'],   W['k3']]
    for i, txt in enumerate(insights):
        r     = 73 + i * 4
        bg_i  = ins_bgs[i % 3]
        acc_i = ins_accs[i % 3]
        for rr in range(r, r + 3):
            for cc in range(2, 21):
                ws.cell(rr, cc).fill = fill(bg_i)
            ws.cell(rr, 2).fill = fill(acc_i)
        try:
            ws.merge_cells(start_row=r, start_column=3, end_row=r+2, end_column=20)
        except Exception:
            pass
        cell_i = ws.cell(r, 3, txt)
        s(cell_i, bg=bg_i, fg=W['txt_dark'], sz=10, h='left', v='center', wrap=True)

    # ── FOOTER
    for c in range(1, 22):
        ws.cell(84, c).fill = fill(W['hdr_bg'])
    ws.merge_cells(start_row=84, start_column=2, end_row=84, end_column=20)
    cell_f = ws.cell(84, 2, 'Dashboard Excel Generator v3.0 · Modern BI · Expert Data Analyst')
    s(cell_f, bg=W['hdr_bg'], fg=W['txt_light'], sz=9, h='center', v='center')

    # Masquer colonnes data
    for c in range(22, 40):
        ws.column_dimensions[get_column_letter(c)].hidden = True
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# PART 8 — FEUILLE ÉVOLUTION
# ══════════════════════════════════════════════════════════════════════════════
def build_evolution_sheet(wb, df, kpis, cm):
    ws = wb.create_sheet('📈 Évolution')
    ws.sheet_view.showGridLines = False

    # Colonnes + hauteurs
    for c in range(1, 12): cw(ws, c, 14)
    rh(ws, 1, 8);  rh(ws, 2, 30);  rh(ws, 3, 8)
    for r in range(4, 40): rh(ws, r, 16)

    # Fond et header
    for r in range(1, 45):
        for c in range(1, 12): ws.cell(r, c).fill = fill(W['bg'])
    for r in [1, 2]:
        for c in range(1, 12): ws.cell(r, c).fill = fill(W['hdr_bg'])
    mg(ws, 2, 1, 2, 11, '📈  ÉVOLUTION MENSUELLE DU CA',
       bg=W['hdr_bg'], fg=W['white'], sz=15, bold=True, h='center', v='center')
    for c in range(1, 12): ws.cell(2, c).border = bottom_border(W['hdr_accent'])

    # Table mensuelle
    monthly = kpis.get('monthly_ca', {})
    sorted_m = sorted(monthly.items())
    hdrs_e = ['Mois/Année', 'CA (€)', 'Croissance %', 'Rang']
    for j, h in enumerate(hdrs_e):
        cell = ws.cell(4, j + 1, h)
        s(cell, bg=W['hdr_bg'], fg=W['white'], sz=10, bold=True, h='center', v='center')

    prev_v = None
    for idx, (m, v) in enumerate(sorted_m, 1):
        r = 4 + idx
        rbg = W['white'] if idx % 2 == 0 else W['muted']
        ws.cell(r, 1, m);   s(ws.cell(r, 1), bg=rbg, fg=W['txt_dark'], sz=10, h='center', v='center')
        ws.cell(r, 2, round(safe_float(v), 2))
        s(ws.cell(r, 2), bg=rbg, fg=W['txt_dark'], sz=10, h='right', v='center',
          nf='# ##0.00 €')
        if prev_v and prev_v != 0:
            growth = (v - prev_v) / abs(prev_v) * 100
            arr, arr_c = trend_arrow(growth)
            ws.cell(r, 3, f"{arr} {fpct(growth)}")
            s(ws.cell(r, 3), bg=rbg, fg=arr_c, sz=10, bold=True, h='center', v='center')
        else:
            ws.cell(r, 3, '—'); s(ws.cell(r, 3), bg=rbg, fg=W['txt_light'], sz=10, h='center', v='center')
        ws.cell(r, 4, idx); s(ws.cell(r, 4), bg=rbg, fg=W['txt_mid'], sz=10, h='center', v='center')
        prev_v = v

    # Graphique Line Chart
    if len(sorted_m) >= 2:
        lc = LineChart()
        lc.title  = "Évolution du CA mensuel"
        lc.style  = 10
        n_rows = len(sorted_m)
        cats  = Reference(ws, min_col=1, min_row=5, max_row=4 + n_rows)
        data  = Reference(ws, min_col=2, min_row=4, max_row=4 + n_rows)
        lc.add_data(data, titles_from_data=True)
        lc.set_categories(cats)
        ser = lc.series[0]
        ser.graphicalProperties.line.solidFill = W['k1']
        ser.graphicalProperties.line.width = 25000
        ser.smooth = True
        lc.width = 18;  lc.height = 10
        ws.add_chart(lc, 'F4')
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# PART 9 — FEUILLE PERFORMANCE
# ══════════════════════════════════════════════════════════════════════════════
def build_performance_sheet(wb, df, kpis, cm):
    ws = wb.create_sheet('🏆 Performance')
    ws.sheet_view.showGridLines = False

    for c in range(1, 16): cw(ws, c, 13)
    rh(ws, 1, 8);  rh(ws, 2, 30);  rh(ws, 3, 8)
    for r in range(4, 50): rh(ws, r, 16)

    for r in range(1, 52):
        for c in range(1, 16): ws.cell(r, c).fill = fill(W['bg'])
    for r in [1, 2]:
        for c in range(1, 16): ws.cell(r, c).fill = fill(W['hdr_bg'])
    mg(ws, 2, 1, 2, 15, '🏆  CLASSEMENT & PERFORMANCE',
       bg=W['hdr_bg'], fg=W['white'], sz=15, bold=True, h='center', v='center')
    for c in range(1, 16): ws.cell(2, c).border = bottom_border(W['hdr_accent'])

    total_ca = kpis.get('ca_total', 1) or 1

    # Table Top Vendeurs (cols 1-5)
    tv = kpis.get('top_vendeurs', {})
    if tv:
        mg(ws, 3, 1, 3, 5, '👤 TOP VENDEURS',
           bg=W['k1'], fg=W['white'], sz=11, bold=True, h='center', v='center')
        for j, h in enumerate(['#', 'Vendeur', 'CA (€)', '% Part', 'Perf.']):
            cell = ws.cell(4, j+1, h)
            s(cell, bg=W['hdr_bg'], fg=W['white'], sz=10, bold=True, h='center', v='center')
        for idx, (k, v) in enumerate(list(tv.items())[:8], 1):
            r   = 4 + idx
            pct = v / total_ca * 100
            rbg = W['white'] if idx % 2 == 0 else W['muted']
            bar = '●' * min(int(pct / 5) + 1, 5)
            for j, cv in enumerate([str(idx), str(k)[:20], feur(v), fpct(pct), bar]):
                cell = ws.cell(r, j+1, cv)
                fg   = W['k2'] if j == 4 else (W['k1'] if j == 0 else W['txt_dark'])
                s(cell, bg=rbg, fg=fg, sz=10, bold=(j==1), h='center' if j!=1 else 'left')
                cell.border = thin_border(W['sep'])

    # Table Top Produits (cols 7-11)
    tp = kpis.get('top_produits', {})
    if tp:
        mg(ws, 3, 7, 3, 11, '📦 TOP PRODUITS',
           bg=W['k2'], fg=W['white'], sz=11, bold=True, h='center', v='center')
        for j, h in enumerate(['#', 'Produit', 'CA (€)', '% Part', 'Perf.']):
            cell = ws.cell(4, 7+j, h)
            s(cell, bg=W['hdr_bg'], fg=W['white'], sz=10, bold=True, h='center', v='center')
        for idx, (k, v) in enumerate(list(tp.items())[:8], 1):
            r   = 4 + idx
            pct = v / total_ca * 100
            rbg = W['white'] if idx % 2 == 0 else W['muted']
            bar = '●' * min(int(pct / 5) + 1, 5)
            for j, cv in enumerate([str(idx), str(k)[:20], feur(v), fpct(pct), bar]):
                cell = ws.cell(r, 7+j, cv)
                fg   = W['k2'] if j == 4 else W['txt_dark']
                s(cell, bg=rbg, fg=fg, sz=10, bold=(j==1), h='center' if j!=1 else 'left')
                cell.border = thin_border(W['sep'])

    # Bar chart Top Vendeurs / Catégories
    src = tv or kpis.get('top_categories', {})
    items_chart = list(src.items())[:8]
    if len(items_chart) >= 2:
        ws.cell(3, 13, 'Segment'); ws.cell(3, 14, 'CA')
        for i, (k, v) in enumerate(items_chart, 1):
            ws.cell(3+i, 13, str(k)[:18])
            ws.cell(3+i, 14, float(v))
        bc = BarChart()
        bc.type    = 'bar'
        bc.title   = "CA par segment"
        bc.style   = 10
        cats_b = Reference(ws, min_col=13, min_row=4, max_row=3+len(items_chart))
        data_b = Reference(ws, min_col=14, min_row=3, max_row=3+len(items_chart))
        bc.add_data(data_b, titles_from_data=True)
        bc.set_categories(cats_b)
        bc.series[0].graphicalProperties.solidFill = W['k3']
        bc.width = 14;  bc.height = 12
        ws.add_chart(bc, 'M4')
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# PART 10 — FEUILLE ANALYSE
# ══════════════════════════════════════════════════════════════════════════════
def build_analyse_sheet(wb, df, kpis, cm):
    ws = wb.create_sheet('🔍 Analyse')
    ws.sheet_view.showGridLines = False

    for c in range(1, 16): cw(ws, c, 14)
    rh(ws, 1, 8);  rh(ws, 2, 30);  rh(ws, 3, 8)
    for r in range(4, 55): rh(ws, r, 16)

    for r in range(1, 57):
        for c in range(1, 16): ws.cell(r, c).fill = fill(W['bg'])
    for r in [1, 2]:
        for c in range(1, 16): ws.cell(r, c).fill = fill(W['hdr_bg'])
    mg(ws, 2, 1, 2, 15, '🔍  ANALYSE DÉTAILLÉE',
       bg=W['hdr_bg'], fg=W['white'], sz=15, bold=True, h='center', v='center')
    for c in range(1, 16): ws.cell(2, c).border = bottom_border(W['hdr_accent'])

    # Section Statuts
    statut_vc = kpis.get('statut_vc', {})
    if statut_vc:
        mg(ws, 3, 1, 3, 5, '📊 RÉPARTITION PAR STATUT',
           bg=W['k4'], fg=W['white'], sz=11, bold=True, h='center', v='center')
        total_s = sum(statut_vc.values())
        STAT_COLORS = {
            'livré': W['k2'], 'livrée': W['k2'], 'delivered': W['k2'],
            'annulé': W['down'], 'annulée': W['down'], 'cancelled': W['down'],
            'en attente': W['k3'], 'en cours': W['k4'],
        }
        for j, h in enumerate(['Statut', 'Nb', '%', 'Barre']):
            cell = ws.cell(4, j+1, h)
            s(cell, bg=W['hdr_bg'], fg=W['white'], sz=10, bold=True, h='center', v='center')
        for idx, (st, cnt) in enumerate(sorted(statut_vc.items(), key=lambda x: -x[1]), 1):
            r   = 4 + idx
            pct = cnt / total_s * 100 if total_s > 0 else 0
            rbg = W['white'] if idx % 2 == 0 else W['muted']
            clr = next((v for k, v in STAT_COLORS.items() if k in str(st).lower()), W['txt_mid'])
            bar = '█' * min(int(pct / 5) + 1, 12)
            for j, cv in enumerate([str(st), str(cnt), fpct(pct), bar]):
                cell = ws.cell(r, j+1, cv)
                fg   = clr if j == 3 else W['txt_dark']
                s(cell, bg=rbg, fg=fg, sz=10, bold=(j==0), h='center' if j != 0 else 'left')
                cell.border = thin_border(W['sep'])

        # Pie chart statuts
        ws.cell(3, 7, 'Statut'); ws.cell(3, 8, 'Nb')
        for i, (st, cnt) in enumerate(statut_vc.items(), 1):
            ws.cell(3+i, 7, str(st)[:15]); ws.cell(3+i, 8, int(cnt))
        pie_s = PieChart()
        pie_s.title = "Répartition Statuts"
        pie_s.style = 10
        cats_s = Reference(ws, min_col=7, min_row=4, max_row=3+len(statut_vc))
        data_s = Reference(ws, min_col=8, min_row=3, max_row=3+len(statut_vc))
        pie_s.add_data(data_s, titles_from_data=True)
        pie_s.set_categories(cats_s)
        pie_s.dataLabels = DataLabelList()
        pie_s.dataLabels.showPercent = True
        for i in range(len(statut_vc)):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = CHART_PAL[i % len(CHART_PAL)]
            pie_s.series[0].dPt.append(pt)
        pie_s.width = 12;  pie_s.height = 10
        ws.add_chart(pie_s, 'G4')

    # Section Remise
    rem = kpis.get('remise_moy')
    if rem is not None:
        row_rem = 4 + len(statut_vc) + 3 if statut_vc else 5
        mg(ws, row_rem, 1, row_rem, 5, '💰 ANALYSE REMISES',
           bg=W['k3'], fg=W['white'], sz=11, bold=True, h='center', v='center')
        for j, (lbl, val) in enumerate([
            ('Remise moyenne', fpct(rem)),
            ('CA total', feur(kpis.get('ca_total', 0))),
            ('Impact remise estimé', feur(kpis.get('ca_total', 0) * rem / 100)),
        ]):
            r = row_rem + 1 + j
            ws.cell(r, 1, lbl); s(ws.cell(r, 1), bg=W['muted'], fg=W['txt_mid'], sz=10, bold=True)
            ws.cell(r, 2, val); s(ws.cell(r, 2), bg=W['white'], fg=W['k3'], sz=12, bold=True, h='center')
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# PART 11 — DONNÉES BRUTES + SOURCE TCD
# ══════════════════════════════════════════════════════════════════════════════
def build_raw_data_sheet(wb, df, cm):
    ws = wb.create_sheet('📋 Données brutes')
    ws.freeze_panes = 'A2'

    stat_col  = cm.get('statut')
    ca_col    = cm.get('ca')
    cols_disp = [c for c in df.columns if c != 'Mois/Année']
    n_cols    = len(cols_disp)

    # En-têtes
    for j, col in enumerate(cols_disp, 1):
        cell = ws.cell(1, j, col)
        s(cell, bg=W['hdr_bg'], fg=W['white'], sz=10, bold=True, h='center', v='center')
        cw(ws, j, max(12, min(len(str(col)) * 1.3, 30)))
    rh(ws, 1, 22)

    STAT_FILLS = {
        'livré': W['k2_l'], 'livrée': W['k2_l'], 'delivered': W['k2_l'],
        'annulé': W['down_l'], 'annulée': W['down_l'], 'cancelled': W['down_l'],
        'en attente': W['k3_l'], 'en cours': W['k4_l'],
    }
    STAT_FG = {
        'livré': W['k2'], 'livrée': W['k2'],
        'annulé': W['down'], 'annulée': W['down'],
        'en attente': W['k3'], 'en cours': W['k4'],
    }

    for i, row in enumerate(df[cols_disp].itertuples(index=False), 2):
        rbg = W['white'] if i % 2 == 0 else W['muted']
        for j, val in enumerate(row, 1):
            col_name = cols_disp[j - 1]
            cell = ws.cell(i, j)
            # Formater valeur — passer par safe_val pour éviter NaN/Inf/NaT
            sv = safe_val(val)
            if sv is None:
                cell.value = ''
            elif isinstance(sv, (int, float)) and not isinstance(sv, bool):
                cell.value = sv
                if ca_col and col_name == ca_col:
                    cell.number_format = '# ##0.00 €'
            else:
                cell.value = str(sv) if str(sv) != 'nan' else ''
            bg = rbg
            fg = W['txt_dark']
            # Couleur statut
            if stat_col and col_name == stat_col:
                val_s = str(val).lower()
                for k, fc in STAT_FILLS.items():
                    if k in val_s:
                        bg = fc
                        fg = STAT_FG.get(k, W['txt_dark'])
                        break
            s(cell, bg=bg, fg=fg, sz=10, h='center', v='center')
        rh(ws, i, 16)

    # Excel Table
    if len(df) > 0 and n_cols > 0:
        try:
            end_col_l = get_column_letter(n_cols)
            tab = Table(displayName='DonneesBrutes',
                        ref=f"A1:{end_col_l}{len(df) + 1}")
            tab.tableStyleInfo = TableStyleInfo(
                name='TableStyleMedium2', showRowStripes=True,
                showFirstColumn=False, showLastColumn=False)
            ws.add_table(tab)
        except Exception as e:
            logger.warning(f"Table Excel: {e}")
    return ws


def build_tcd_source_sheet(wb, df, cm):
    ws = wb.create_sheet('📝 Source TCD')
    ws.freeze_panes = 'A2'

    # Colonnes pertinentes pour TCD
    tcd_roles = ['date', 'client', 'produit', 'categorie', 'marque',
                 'quantite', 'remise', 'ca', 'tva', 'statut', 'vendeur', 'region']
    tcd_cols = [cm[r] for r in tcd_roles if cm.get(r) and cm[r] in df.columns]
    if 'Mois/Année' in df.columns:
        tcd_cols = ['Mois/Année'] + [c for c in tcd_cols if c != 'Mois/Année']
    if not tcd_cols:
        tcd_cols = list(df.columns)

    for j, col in enumerate(tcd_cols, 1):
        cell = ws.cell(1, j, col)
        s(cell, bg=W['hdr_bg'], fg=W['white'], sz=10, bold=True, h='center', v='center')
        cw(ws, j, 16)
    rh(ws, 1, 22)

    for i, row in enumerate(df[tcd_cols].itertuples(index=False), 2):
        rbg = W['white'] if i % 2 == 0 else W['muted']
        for j, val in enumerate(row, 1):
            cell = ws.cell(i, j)
            sv = safe_val(val)
            cell.value = sv if (sv is not None and str(sv) != 'nan') else ''
            s(cell, bg=rbg, fg=W['txt_dark'], sz=10, h='center', v='center')
        rh(ws, i, 15)

    if len(df) > 0 and len(tcd_cols) > 0:
        try:
            tab = Table(displayName='SourceTCD',
                        ref=f"A1:{get_column_letter(len(tcd_cols))}{len(df)+1}")
            tab.tableStyleInfo = TableStyleInfo(
                name='TableStyleLight1', showRowStripes=True)
            ws.add_table(tab)
        except Exception as e:
            logger.warning(f"Source TCD Table: {e}")
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# PART 12 — ORCHESTRATEUR PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
def generate_excel_dashboard(file_bytes, filename, user_email=''):
    """Pipeline complet : load → clean → detect → kpis → build → bytes."""
    logger.info(f"Génération dashboard: {filename} ({len(file_bytes)} bytes)")

    # 1. Chargement
    df = load_dataframe(file_bytes, filename)
    if df is None or len(df) == 0:
        raise ValueError("Fichier vide ou illisible")
    logger.info(f"DataFrame chargé: {df.shape}")

    # 2. Nettoyage
    df = clean_dataframe(df)

    # 3. Détection colonnes
    cm = detect_columns(df)

    # 4. KPIs avancés
    kpis = compute_advanced_kpis(df, cm)

    # 5. AI Insights
    insights = generate_insights(df, kpis, cm)

    # 6. Construction workbook
    wb = Workbook()
    # Supprimer feuille par défaut
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    build_modern_dashboard(wb, df, kpis, cm, insights)
    build_evolution_sheet(wb, df, kpis, cm)
    build_performance_sheet(wb, df, kpis, cm)
    build_analyse_sheet(wb, df, kpis, cm)
    build_raw_data_sheet(wb, df, cm)
    build_tcd_source_sheet(wb, df, cm)

    # 7. Sérialisation en bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    excel_bytes = buf.read()
    logger.info(f"Excel généré: {len(excel_bytes)} bytes, {len(wb.sheetnames)} onglets")

    return {
        'status':        'success',
        'excel_base64':  base64.b64encode(excel_bytes).decode('utf-8'),
        'filename':      f"Dashboard_{filename.rsplit('.', 1)[0]}.xlsx",
        'kpis': {
            'total_rows':       kpis['n_rows'],
            'ca_total':         round(kpis.get('ca_total', 0), 2),
            'n_clients':        kpis.get('n_clients', 0),
            'panier_moyen':     round(kpis.get('panier_moy', 0), 2),
            'taux_livraison':   round(kpis.get('taux_livraison') or 0, 1),
            'taux_annulation':  round(kpis.get('taux_annulation') or 0, 1),
            'mom_growth':       round(kpis.get('mom_growth') or 0, 1),
            'n_onglets':        len(wb.sheetnames),
        },
        'insights': insights,
        'email':    user_email,
    }


# ══════════════════════════════════════════════════════════════════════════════
# PART 13 — ROUTES FLASK
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/health', methods=['GET'])
def health():
    return jsonify({
        'status':    'ok',
        'version':   '3.0.0',
        'timestamp': datetime.now().isoformat(),
    })


@app.route('/generate-dashboard', methods=['POST'])
def generate_dashboard():
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'status': 'error', 'message': 'Corps JSON manquant'}), 400

        # Récupérer fichier (base64 ou multipart)
        file_data = data.get('file_data', '')
        filename  = data.get('filename', 'fichier.xlsx')
        email     = data.get('email', '')

        if not file_data:
            return jsonify({'status': 'error', 'message': 'Champ file_data manquant'}), 400

        # Décodage base64
        try:
            file_bytes = base64.b64decode(file_data)
        except Exception:
            return jsonify({'status': 'error', 'message': 'file_data base64 invalide'}), 400

        result = generate_excel_dashboard(file_bytes, filename, email)
        return jsonify(result)

    except ValueError as ve:
        import traceback as _tb
        tb_str = _tb.format_exc()
        logger.error(f"ValueError: {ve}\n{tb_str}")
        return jsonify({'status': 'error', 'message': str(ve), 'traceback': tb_str}), 422
    except Exception as e:
        import traceback as _tb
        tb_str = _tb.format_exc()
        logger.error(f"Erreur inattendue: {e}\n{tb_str}", exc_info=False)
        return jsonify({'status': 'error', 'message': f'Erreur serveur: {str(e)}', 'traceback': tb_str}), 500


@app.route('/generate-from-upload', methods=['POST'])
def generate_from_upload():
    """Route multipart/form-data pour fichiers volumineux."""
    try:
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'Fichier manquant'}), 400
        f         = request.files['file']
        email     = request.form.get('email', '')
        file_bytes = f.read()
        result    = generate_excel_dashboard(file_bytes, f.filename or 'upload.xlsx', email)
        return jsonify(result)
    except Exception as e:
        logger.error(f"generate-from-upload: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
